import json
from openpyxl import Workbook

def writeDataToExcel(data):
    wb = Workbook()
    ws = wb.active

    i = 0
    for number in data:
        i += 1
        ws.cell(row=i, column=1).value = number

        values = data.get(number)
        for j in range(len(values)):
            ws.cell(row=i, column=j+2).value = values[j]

    wb.save('sample.xlsx')

if __name__ == '__main__':
    with open('student.txt', 'r', encoding='utf-8') as f:
        data = json.load(f)
        writeDataToExcel(data)
