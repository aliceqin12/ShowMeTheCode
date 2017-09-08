from openpyxl import load_workbook
from xml.dom.minidom import Document
from collections import OrderedDict
import json
import html


comment = '''
            学生信息表
            "id" : [名字, 数学, 语文, 英文]
        '''

def readExcelDataToDict(excelpath):
    wb = load_workbook(filename=excelpath)
    ws = wb.get_sheet_by_name('Sheet')
    student_dict = OrderedDict()

    for i in range(3):
        id = ws.cell(row=i+1, column=1).value
        student_values = []
        for j in range(4):
            value = ws.cell(row=i+1, column=j+2).value
            student_values.append(value)
        student_dict[id] = student_values

    return student_dict

def writeDataToXml(student_dict, xmlpath):
    doc = Document()
    root_node = doc.createElement('root')
    doc.appendChild(root_node)

    stu_node = doc.createElement('students')
    root_node.appendChild(stu_node)

    note_node = doc.createComment(comment)
    stu_node.appendChild(note_node)

    dict_node = doc.createTextNode(json.dumps(student_dict, ensure_ascii=False))
    stu_node.appendChild(dict_node)

    with open(xmlpath, "w", encoding='utf8') as file:
        transform = html.unescape(doc.toprettyxml())
        file.write(transform)

if __name__ == '__main__':
    excelpath = 'student.xlsx'
    xmlpath = 'student.xml'
    student_dict = readExcelDataToDict(excelpath)
    writeDataToXml(student_dict, xmlpath)